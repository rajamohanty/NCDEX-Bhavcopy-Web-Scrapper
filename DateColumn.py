#!/usr/bin/env python
# coding: utf-8

# In[ ]:


def DateColumn(path):
    '''
    Creates a date column for each bhavcopy excel files.
    
    Input Parameters:
    path -> Put in the path of the directory where all the Bhavcopy xlsx files are located.
    
    '''
    import os
    import xlrd
    import openpyxl

    os.chdir(path)

    file_names = os.listdir(path)

    ########################################### Get Number of Rows and Columns

    total_columns = []
    total_rows = []
    for file in file_names:
        workbook = xlrd.open_workbook(file)
        worksheet = workbook.sheet_by_index(0)
        total_rows.append(worksheet.nrows)
        total_columns.append(worksheet.ncols)

    header = "Date"

    current_date =[]
    for f in os.listdir():
        f_name, f_ext = os.path.splitext(f)
        current_date.append(f_name)


    ########################################## Write Date Column in each files
    for (file,c,cd,r) in zip(file_names, total_columns,current_date,total_rows):
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        worksheet.cell(row = 1, column = c+1).value=header
        for rw in range(2,r+1):
            worksheet.cell(row = rw, column = c+1).value=cd
        workbook.save(file)    

